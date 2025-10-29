"""
Excel File Parser
Extracts product data from Excel spreadsheets with support for multiple formats
"""

import os
from pathlib import Path
from typing import Dict, List, Optional, Any
import logging
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

logger = logging.getLogger(__name__)


@dataclass
class ProductRow:
    """Single product row extracted from Excel"""
    sku: str
    description: str
    quantity: int
    unit_price: float
    total_price: Optional[float] = None
    category: Optional[str] = None
    row_number: int = 0
    raw_data: Dict[str, Any] = None


@dataclass
class ExcelSheetData:
    """Data extracted from a single Excel sheet"""
    sheet_name: str
    headers: List[str]
    products: List[ProductRow]
    raw_data: List[Dict]


class ExcelParser:
    """Parse Excel files to extract product information"""
    
    def __init__(self):
        if openpyxl is None and pd is None:
            logger.warning("No Excel parsing library available")
    
    def parse_excel(self, filepath: str, sheet_names: Optional[List[str]] = None) -> Dict[str, ExcelSheetData]:
        """
        Parse Excel file and extract product data
        
        Args:
            filepath: Path to Excel file
            sheet_names: Specific sheets to parse (None = all sheets)
            
        Returns:
            Dictionary mapping sheet names to ExcelSheetData
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel file not found: {filepath}")
        
        try:
            if pd:
                return self._parse_with_pandas(filepath, sheet_names)
            elif openpyxl:
                return self._parse_with_openpyxl(filepath, sheet_names)
            else:
                raise ImportError("No Excel parsing library available")
        except Exception as e:
            logger.error(f"Error parsing Excel file {filepath}: {e}")
            raise
    
    def _parse_with_pandas(self, filepath: str, sheet_names: Optional[List[str]] = None) -> Dict[str, ExcelSheetData]:
        """Parse using pandas (handles .xlsx and .xls)"""
        all_sheets = {}
        
        try:
            excel_file = pd.ExcelFile(filepath)
            sheets_to_process = sheet_names if sheet_names else excel_file.sheet_names
            
            for sheet_name in sheets_to_process:
                try:
                    # Try to read without headers first to detect header row
                    df_no_header = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    # Find header row by looking for common column names
                    header_row = self._find_header_row(df_no_header)
                    
                    if header_row is not None:
                        # Read with header at found row
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
                        # Remove rows above header if there are any meaningful differences
                        if header_row > 0:
                            df = df.dropna(how='all')  # Remove completely empty rows
                    else:
                        # Fallback to default reading
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    sheet_data = self._extract_products_from_dataframe(df, sheet_name)
                    all_sheets[sheet_name] = sheet_data
                except Exception as e:
                    logger.warning(f"Error processing sheet {sheet_name}: {e}")
                    continue
        
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
        
        return all_sheets
    
    def _find_header_row(self, df: pd.DataFrame, max_rows_to_check: int = 20) -> Optional[int]:
        """
        Find the row that contains headers by looking for common column name patterns
        
        Args:
            df: DataFrame read without headers
            max_rows_to_check: Maximum number of rows to check
            
        Returns:
            Row index if found, None otherwise
        """
        header_keywords = [
            'product', 'sku', 'part', 'item', 'qty', 'quantity', 'desc', 'description',
            'price', 'cost', 'unit', 'total', 'part number', 'item number',
            'מק"ט', 'תיאור', 'כמות', 'מחיר', 'מוצר'
        ]
        
        # Check first N rows
        for row_idx in range(min(max_rows_to_check, len(df))):
            row_values = [str(val).lower().strip() for val in df.iloc[row_idx].values if pd.notna(val)]
            
            # Count how many header keywords we find in this row
            keyword_matches = sum(1 for val in row_values 
                                 if any(keyword in val for keyword in header_keywords))
            
            # If we find at least 2-3 keywords, this is likely a header row
            if keyword_matches >= 2:
                logger.debug(f"Found potential header row at index {row_idx}: {row_values[:5]}")
                return row_idx
        
        return None
    
    def _parse_with_openpyxl(self, filepath: str, sheet_names: Optional[List[str]] = None) -> Dict[str, ExcelSheetData]:
        """Parse using openpyxl (handles .xlsx only)"""
        all_sheets = {}
        
        try:
            workbook = load_workbook(filepath, data_only=True)
            sheets_to_process = sheet_names if sheet_names else workbook.sheetnames
            
            for sheet_name in sheets_to_process:
                try:
                    worksheet = workbook[sheet_name]
                    df = pd.DataFrame(worksheet.values)
                    
                    # Set first row as headers
                    if len(df) > 0:
                        df.columns = df.iloc[0]
                        df = df[1:].reset_index(drop=True)
                    
                    sheet_data = self._extract_products_from_dataframe(df, sheet_name)
                    all_sheets[sheet_name] = sheet_data
                except Exception as e:
                    logger.warning(f"Error processing sheet {sheet_name}: {e}")
                    continue
        
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
        
        return all_sheets
    
    def _extract_products_from_dataframe(self, df: pd.DataFrame, sheet_name: str) -> ExcelSheetData:
        """
        Extract product data from pandas DataFrame
        
        Attempts to identify columns containing SKU, description, quantity, price
        """
        if df.empty:
            return ExcelSheetData(
                sheet_name=sheet_name,
                headers=[],
                products=[],
                raw_data=[]
            )
        
        headers = list(df.columns)
        products = []
        raw_data = df.to_dict('records')
        
        # Try to identify relevant columns
        sku_col = self._find_column(df, ['sku', 'part number', 'part#', 'item', 'product code', 'product', 'מק"ט', 'מוצר'])
        desc_col = self._find_column(df, ['description', 'item description', 'desc', 'תיאור', 'מוצר'])
        qty_col = self._find_column(df, ['quantity', 'qty', 'amount', 'כמות', 'מספר'])
        price_col = self._find_column(df, ['price', 'unit price', 'cost', 'unit cost', 'מחיר', 'מחיר יחידה'])
        total_col = self._find_column(df, ['total', 'line total', 'total price', 'extended', 'סה"כ', 'סה"כ שורה'])
        
        # Log what was found
        logger.info(f"Column detection - SKU={sku_col}, Desc={desc_col}, Qty={qty_col}, Price={price_col}, Total={total_col}")
        
        # Require at least SKU, Description, and Quantity. Price is optional (may calculate later)
        if not all([sku_col, desc_col, qty_col]):
            logger.warning(f"Could not identify required columns (SKU, Description, Quantity) in sheet {sheet_name}")
            logger.info(f"Found columns: SKU={sku_col}, Desc={desc_col}, Qty={qty_col}, Price={price_col}")
        
        # First pass: collect all rows and detect prices in description column
        # Prices are often in column C (description) on rows below products
        row_data_list = []
        for idx, row in df.iterrows():
            row_data_list.append((idx, row, df.columns))
        
        # Second pass: extract products and find their prices
        # Prices might be in the same row, or in the row below (in description column)
        for idx, (row_idx, row, columns) in enumerate(row_data_list):
            try:
                sku = str(row[sku_col]).strip() if sku_col and sku_col in row else ""
                description = str(row[desc_col]).strip() if desc_col and desc_col in row else ""
                
                # Skip header rows, empty rows, and price rows (rows with "includes", "total", etc. but no real product)
                # Price rows typically have labels like "includes:..." or "Total" in PRODUCT column
                is_price_row = False
                if sku:
                    sku_lower = sku.lower()
                    is_price_row = (
                        'includes' in sku_lower or 
                        sku_lower == 'total' or 
                        sku_lower.startswith('total') or
                        'support' in sku_lower and 'delivery' in sku_lower
                    )
                
                if not sku or sku.lower() in ['sku', 'part number', 'מק"ט', 'product', ''] or is_price_row:
                    continue
                
                # Parse quantity
                quantity = self._safe_int(row.get(qty_col) if qty_col else None, 1)
                
                # Parse prices - flexible multi-strategy approach for different Excel formats
                unit_price = self._safe_float(row.get(price_col) if price_col else None, 0.0)
                total_price = self._safe_float(row.get(total_col) if total_col else None, None)
                
                # Strategy 1: Price in dedicated price/total columns (same row)
                if unit_price == 0.0 and total_price is None:
                    # Strategy 2: Price in description column (same row)
                    if desc_col:
                        desc_val = row.get(desc_col)
                        if isinstance(desc_val, (int, float)) and desc_val > 100:
                            total_price = self._safe_float(desc_val, None)
                            if total_price and quantity > 0:
                                unit_price = total_price / quantity
                                logger.debug(f"Found price ${total_price:.2f} for {sku} in description column (same row)")
                    
                    # Strategy 3: Check rows BELOW for price (prices often appear below products)
                    # Check up to 3-5 rows ahead, but stop if we hit another product
                    if unit_price == 0.0 and total_price is None:
                        for offset in range(1, min(6, len(row_data_list) - idx)):  # Check up to 5 rows below
                            if idx + offset >= len(row_data_list):
                                break
                            
                            check_row_idx, check_row, _ = row_data_list[idx + offset]
                            check_row_sku = str(check_row.get(sku_col, '')).strip() if sku_col else ""
                            
                            # If we hit another real product before finding price, stop searching
                            if check_row_sku and check_row_sku.lower() not in ['', 'total', 'includes', 'support']:
                                # Check if this is actually a price row disguised as a product
                                check_desc = check_row.get(desc_col) if desc_col else None
                                if not (isinstance(check_desc, (int, float)) and check_desc > 100):
                                    # This is a real product, stop looking
                                    break
                            
                            check_row_label = check_row_sku
                            check_desc_val = check_row.get(desc_col) if desc_col else None
                            check_price_col_val = check_row.get(price_col) if price_col else None
                            check_total_col_val = check_row.get(total_col) if total_col else None
                            
                            # Check all possible price locations in this row
                            found_price = None
                            is_price_row = False
                            
                            # Check description column for price (most common in this Excel format)
                            if check_desc_val:
                                price_candidate = self._safe_float(check_desc_val, None)
                                if price_candidate and price_candidate > 100:
                                    found_price = price_candidate
                                    # Determine if this looks like a price row
                                    is_price_row = (
                                        ('includes' in check_row_label.lower()) or
                                        (check_row_label.lower() == 'total') or
                                        (check_row_label.lower().startswith('total')) or
                                        (not check_row_label or check_row_label == '' or pd.isna(check_row.get(sku_col)) if sku_col else True)
                                    )
                            
                            # Check price/total columns if not found in description
                            if not found_price:
                                if check_price_col_val:
                                    price_candidate = self._safe_float(check_price_col_val, None)
                                    if price_candidate and price_candidate > 100:
                                        found_price = price_candidate
                                        is_price_row = True
                                
                                if not found_price and check_total_col_val:
                                    price_candidate = self._safe_float(check_total_col_val, None)
                                    if price_candidate and price_candidate > 100:
                                        found_price = price_candidate
                                        is_price_row = True
                            
                            # If we found a price and it looks like a price row, use it
                            if found_price and is_price_row:
                                total_price = found_price
                                if quantity > 0:
                                    unit_price = total_price / quantity
                                logger.info(f"Found price ${total_price:.2f} for {sku} {offset} row(s) below (row label: '{check_row_label}')")
                                break
                    
                    # Strategy 4: Check all columns in same row for price-like values
                    if unit_price == 0.0 and total_price is None:
                        for col_name in df.columns:
                            if col_name in [sku_col, desc_col, qty_col]:
                                continue
                            col_val = row.get(col_name)
                            if isinstance(col_val, (int, float)) and col_val > 100:
                                col_lower = str(col_name).lower()
                                if any(keyword in col_lower for keyword in ['price', 'cost', 'amount', 'total', '$', 'מחיר']):
                                    total_price = self._safe_float(col_val, None)
                                    if total_price and quantity > 0:
                                        unit_price = total_price / quantity
                                        logger.debug(f"Found price ${total_price:.2f} for {sku} in column '{col_name}' (same row)")
                                        break
                
                # Calculate total if not provided
                if total_price is None and unit_price and quantity:
                    total_price = unit_price * quantity
                
                # Allow products without price (price can be added later or extracted from email)
                # Minimum requirement: SKU and Description
                if sku and description:
                    products.append(ProductRow(
                        sku=sku,
                        description=description,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price,
                        row_number=idx + 2,  # +1 for 0-index, +1 for header
                        raw_data=row.to_dict()
                    ))
            
            except Exception as e:
                logger.warning(f"Error processing row {idx} in sheet {sheet_name}: {e}")
                continue
        
        return ExcelSheetData(
            sheet_name=sheet_name,
            headers=headers,
            products=products,
            raw_data=raw_data
        )
    
    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """Find column by matching possible names (case-insensitive)"""
        columns_lower = {col.lower(): col for col in df.columns}
        
        for name in possible_names:
            name_lower = name.lower()
            # Exact match
            if name_lower in columns_lower:
                return columns_lower[name_lower]
            # Partial match
            for col in df.columns:
                if name_lower in col.lower() or col.lower() in name_lower:
                    return col
        
        return None
    
    def _safe_int(self, value: Any, default: int = 0) -> int:
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            if isinstance(value, str):
                # Remove commas, spaces
                value = value.replace(',', '').replace(' ', '')
            return int(float(value))
        except (ValueError, TypeError):
            return default
    
    def _safe_float(self, value: Any, default: Optional[float] = None) -> Optional[float]:
        """Safely convert value to float"""
        if value is None:
            return default
        try:
            if isinstance(value, str):
                # Remove currency symbols, commas
                value = value.replace('$', '').replace('₪', '').replace(',', '').replace(' ', '')
            result = float(value)
            return result if result >= 0 else default
        except (ValueError, TypeError):
            return default
    
    def merge_sheets(self, sheets_data: Dict[str, ExcelSheetData]) -> List[ProductRow]:
        """
        Merge products from multiple sheets into single list
        
        Args:
            sheets_data: Dictionary of sheet data
            
        Returns:
            Combined list of ProductRow objects
        """
        all_products = []
        
        for sheet_data in sheets_data.values():
            all_products.extend(sheet_data.products)
        
        return all_products

