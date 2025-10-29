"""
Quote Generator
Creates formatted Excel quote documents with pricing
"""

import os
from pathlib import Path
from typing import List, Optional, Dict
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None

from .pricing import PricedProduct, QuoteSummary

logger = logging.getLogger(__name__)


class QuoteGenerator:
    """Generate formatted Excel quote documents"""
    
    def __init__(self, config: Dict):
        """
        Initialize quote generator
        
        Args:
            config: Configuration dictionary
        """
        self.config = config
        self.company_name = config.get('company', {}).get('name', 'Dayo Tech')
        self.company_email = config.get('company', {}).get('email', 'info@dayo-tech.com')
        
        if Workbook is None:
            logger.error("openpyxl not available. Cannot generate Excel quotes.")
    
    def generate_quote(self, 
                      priced_products: List[PricedProduct],
                      summary: QuoteSummary,
                      output_path: str,
                      customer_name: Optional[str] = None,
                      quote_number: Optional[str] = None,
                      currency: str = "USD") -> str:
        """
        Generate formatted Excel quote
        
        Args:
            priced_products: List of PricedProduct objects
            summary: QuoteSummary object
            output_path: Path to save quote file
            customer_name: Customer name for header
            quote_number: Quote number/ID
            currency: Currency symbol/code
            
        Returns:
            Path to generated quote file
        """
        if Workbook is None:
            raise ImportError("openpyxl is required for quote generation")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Quote"
        
        # Generate quote number if not provided
        if quote_number is None:
            quote_number = f"QT-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Header section
        row = 1
        self._add_header(ws, row, customer_name, quote_number, currency)
        row += 6
        
        # Column headers
        headers = [
            "SKU",
            "Description (English)",
            "Description (Hebrew)",
            "Quantity",
            f"Unit Price ({currency})",
            "Margin %",
            f"Unit Selling Price ({currency})",
            f"Total Line ({currency})"
        ]
        self._add_table_header(ws, row, headers)
        row += 1
        
        # Product rows
        for product in priced_products:
            self._add_product_row(ws, row, product, currency)
            row += 1
        
        # Summary section
        row += 1
        self._add_summary(ws, row, summary, currency)
        
        # Footer
        row += 3
        self._add_footer(ws, row)
        
        # Auto-size columns
        self._auto_size_columns(ws, len(headers))
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
        wb.save(output_path)
        
        logger.info(f"Generated quote saved to {output_path}")
        return output_path
    
    def _add_header(self, ws, row: int, customer_name: Optional[str], quote_number: str, currency: str):
        """Add quote header"""
        # Company name
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_name
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "QUOTE"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Quote Number: {quote_number}"
        cell.font = Font(size=12)
        cell.alignment = Alignment(horizontal='center')
        
        if customer_name:
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = f"Customer: {customer_name}"
            cell.font = Font(size=12)
        
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(size=11)
    
    def _add_table_header(self, ws, row: int, headers: List[str]):
        """Add table header row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    
    def _add_product_row(self, ws, row: int, product: PricedProduct, currency: str):
        """Add product data row"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # For now, use same description for English/Hebrew (can be enhanced with translation)
        description = product.description
        
        values = [
            product.sku,
            description,  # English (can be split if bilingual)
            description,  # Hebrew (can be translated)
            product.quantity,
            product.vendor_unit_price,  # Vendor unit price
            product.margin_percent,
            product.selling_unit_price,
            product.selling_total
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Format numbers
            if isinstance(value, (int, float)) and col >= 4:
                if col in [4, 6, 7, 8]:  # Price columns
                    cell.number_format = '#,##0.00'
                elif col == 5:  # Margin %
                    cell.number_format = '0.00%'
            
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True) if col == 2 or col == 3 else Alignment(vertical='center')
    
    def _add_summary(self, ws, row: int, summary: QuoteSummary, currency: str):
        """Add summary totals"""
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        summary_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # Summary header
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "SUMMARY"
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = border
        
        row += 1
        
        # Total vendor cost
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total Vendor Cost ({currency}):"
        cell.font = Font(bold=True)
        
        cell = ws[f'E{row}']
        cell.value = summary.total_vendor_cost
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        
        row += 1
        
        # Total margin
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total Margin ({currency}):"
        cell.font = Font(bold=True)
        
        cell = ws[f'E{row}']
        cell.value = summary.total_margin
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        
        row += 1
        
        # Average margin %
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "Average Margin %:"
        cell.font = Font(bold=True)
        
        cell = ws[f'E{row}']
        cell.value = summary.margin_percent_avg / 100
        cell.number_format = '0.00%'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        
        row += 1
        
        # Grand total
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"GRAND TOTAL ({currency}):"
        cell.font = Font(bold=True, size=12)
        cell.fill = summary_fill
        
        cell = ws[f'E{row}']
        cell.value = summary.total_selling_price
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, size=12)
        cell.fill = summary_fill
        cell.alignment = Alignment(horizontal='right')
    
    def _add_footer(self, ws, row: int):
        """Add footer with terms and contact"""
        footer_text = f"""
Terms and Conditions:
- Prices are valid for 30 days
- Payment terms: Net 30
- Shipping costs not included unless specified

Contact:
{self.company_email}
"""
        
        ws.merge_cells(f'A{row}:H{row + 4}')
        cell = ws[f'A{row}']
        cell.value = footer_text
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    def _auto_size_columns(self, ws, num_columns: int):
        """Auto-size columns for better readability"""
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            # Set column width (with some limits)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_project_quote(self,
                              priced_products: List[PricedProduct],
                              summary: QuoteSummary,
                              output_path: str,
                              customer_name: Optional[str] = None,
                              quote_number: Optional[str] = None,
                              vendor_grouping: Optional[Dict[str, List[PricedProduct]]] = None,
                              currency: str = "USD") -> str:
        """
        Generate project-style quote with vendor sheets and proposal sheet
        (Based on project example.xlsx structure)
        
        Args:
            priced_products: List of PricedProduct objects
            summary: QuoteSummary object
            output_path: Path to save quote file
            customer_name: Customer name
            quote_number: Quote number/ID
            vendor_grouping: Dict mapping vendor/category names to products
            currency: Currency symbol/code
            
        Returns:
            Path to generated quote file
        """
        if Workbook is None:
            raise ImportError("openpyxl is required for quote generation")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Group products by vendor/category if not provided
        if vendor_grouping is None:
            vendor_grouping = self._group_products_by_vendor(priced_products)
        
        # Create vendor sheets
        vendor_sheets = {}
        for vendor_name, products in vendor_grouping.items():
            sheet_name = self._sanitize_sheet_name(vendor_name)
            ws = self._create_vendor_sheet(wb, sheet_name, products, currency)
            vendor_sheets[vendor_name] = ws
        
        # Create proposal sheet (consolidates all products)
        proposal_sheet = self._create_proposal_sheet(wb, priced_products, customer_name, quote_number, currency)
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
        wb.save(output_path)
        
        logger.info(f"Generated project quote with {len(vendor_grouping)} vendor sheets saved to {output_path}")
        return output_path
    
    def _group_products_by_vendor(self, products: List[PricedProduct]) -> Dict[str, List[PricedProduct]]:
        """
        Group products by vendor/category
        
        For now, groups by product category or source metadata
        Can be enhanced with vendor detection logic
        """
        grouping = {}
        
        for product in products:
            # Try to get vendor/category from metadata
            vendor = None
            if product.raw_product:
                vendor = product.raw_product.metadata.get('vendor') or product.raw_product.metadata.get('category')
            
            # Use category or default
            if not vendor:
                vendor = product.category or "General"
            
            if vendor not in grouping:
                grouping[vendor] = []
            grouping[vendor].append(product)
        
        return grouping
    
    def _sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """Sanitize sheet name for Excel (max 31 chars, no special chars)"""
        # Remove invalid chars: / \ ? * [ ]
        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Truncate if needed
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]
        
        return sanitized
    
    def _create_vendor_sheet(self, wb, sheet_name: str, products: List[PricedProduct], currency: str):
        """
        Create vendor sheet with profit/margin columns
        Format: PRODUCT | DESCRIPTION | QTY | Price | Total price | Sale | Total Sale | mergin
        """
        ws = wb.create_sheet(title=sheet_name)
        
        # Header row
        headers = ['PRODUCT', 'DESCRIPTION', 'QTY', 'Price', 'Total price', 'Sale', 'Total Sale', 'mergin']
        header_row = 1
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Product rows
        row = header_row + 1
        for product in products:
            # Calculate values
            vendor_price = product.vendor_unit_price
            total_price = product.vendor_total if product.vendor_total else (vendor_price * product.quantity)
            sale_unit = product.selling_unit_price
            total_sale = product.selling_total
            margin = product.margin_percent
            
            values = [
                product.sku,
                product.description[:200] if product.description else '',  # Truncate long descriptions
                product.quantity,
                vendor_price,
                total_price,
                sale_unit,
                total_sale,
                margin
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Format numbers
                if isinstance(value, (int, float)):
                    if col in [4, 5, 6, 7]:  # Price columns
                        cell.number_format = '#,##0.00'
                    elif col == 8:  # Margin
                        cell.number_format = '0.00%'
                
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(vertical='top', wrap_text=True) if col == 2 else Alignment(vertical='center')
            
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def _create_proposal_sheet(self, wb, products: List[PricedProduct], customer_name: Optional[str], quote_number: Optional[str], currency: str):
        """
        Create proposal sheet (consolidates all products for customer)
        Format: Product | Description | Qty | Price | Total | Mergin | Profit %
        """
        ws = wb.create_sheet(title="proposal")
        
        # Header at row 3 (as in example)
        header_row = 3
        headers = ['Product', 'Description', 'Qty', 'Price', 'Total', 'Mergin', 'Profit %']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col + 1)  # Start at column B (skip first column)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Product rows
        row = header_row + 1
        for product in products:
            values = [
                product.sku,
                product.description[:200] if product.description else '',
                product.quantity,
                product.selling_unit_price,  # Selling price in proposal
                product.selling_total,
                product.margin_amount,
                product.margin_percent
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col + 1)  # Start at column B
                cell.value = value
                
                # Format numbers
                if isinstance(value, (int, float)):
                    if col in [4, 5, 6]:  # Price/margin columns
                        cell.number_format = '#,##0.00'
                    elif col == 7:  # Profit %
                        cell.number_format = '0.00%'
                
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(vertical='top', wrap_text=True) if col == 2 else Alignment(vertical='center')
            
            row += 1
        
        # Auto-size columns
        for col in range(2, len(headers) + 2):  # Columns B-H
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws

